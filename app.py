import os
import math
from functools import wraps
from datetime import datetime
from zoneinfo import ZoneInfo

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import (
    LoginManager, UserMixin, login_user, logout_user,
    login_required, current_user
)
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

import config
from models import get_db, init_db
import sync_data

TZ_COL = ZoneInfo('America/Bogota')

app = Flask(__name__)
app.secret_key = config.SECRET_KEY
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

init_db()

login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Debes iniciar sesión para acceder.'
login_manager.login_message_category = 'error'


class User(UserMixin):
    def __init__(self, row):
        self.id = row['id']
        self.username = row['username']
        self.nombre_completo = row['nombre_completo']
        self.rol = row['rol']
        self.activo = row['activo']

    def get_id(self):
        return str(self.id)


@login_manager.user_loader
def load_user(user_id):
    conn = get_db()
    row = conn.execute(
        "SELECT * FROM usuarios WHERE id = ? AND activo = 1", (user_id,)
    ).fetchone()
    conn.close()
    return User(row) if row else None


def admin_required(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if current_user.rol not in ('admin', 'superadmin'):
            flash('Acceso restringido a administradores.', 'error')
            return redirect(url_for('dashboard_redirect'))
        return f(*args, **kwargs)
    return decorated


def cio_required(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if current_user.rol != 'cio':
            flash('Acceso restringido a CIO.', 'error')
            return redirect(url_for('dashboard_redirect'))
        return f(*args, **kwargs)
    return decorated


def tecnico_required(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if current_user.rol not in ('tecnico', 'admin', 'superadmin'):
            flash('Acceso restringido.', 'error')
            return redirect(url_for('dashboard_redirect'))
        return f(*args, **kwargs)
    return decorated


@app.context_processor
def inject_no_reportadas():
    if not current_user.is_authenticated:
        return {}
    if current_user.rol not in ('admin', 'superadmin', 'cio'):
        return {}
    try:
        conn = get_db()
        c = conn.execute(
            "SELECT COUNT(*) AS c FROM ejecuciones_no_reportadas WHERE estado = 'pendiente'"
        ).fetchone()['c']
        conn.close()
        return {'no_reportadas_pendientes': c}
    except Exception:
        return {'no_reportadas_pendientes': 0}


def _log_actividad(conn, usuario_id, accion_tipo, detalle):
    conn.execute(
        """INSERT INTO log_actividad (usuario_id, accion_tipo, detalle, ip_address, timestamp)
           VALUES (?, ?, ?, ?, ?)""",
        (usuario_id, accion_tipo, detalle, request.remote_addr, datetime.now(TZ_COL).isoformat())
    )


def _current_sync_id(conn):
    row = conn.execute("SELECT MAX(sync_id) AS sync_id FROM equipos").fetchone()
    return row['sync_id'] if row and row['sync_id'] else None


_MESES_ES = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']

def _format_fecha_actualizacion(fecha_str):
    if not fecha_str:
        return None
    s = str(fecha_str).strip()
    dt = None
    for fmt in ('%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M'):
        try:
            dt = datetime.strptime(s, fmt)
            break
        except ValueError:
            pass
    if dt is None:
        try:
            dt = datetime.strptime(s[:19], '%Y-%m-%dT%H:%M:%S')
        except ValueError:
            return s
    mes  = _MESES_ES[dt.month - 1]
    ampm = 'a.m.' if dt.hour < 12 else 'p.m.'
    return f"{dt.day:02d}/{mes}/{dt.year} — {dt.strftime('%I:%M')} {ampm}"


_ALLOWED_EXCEL = {'xlsx', 'xls'}

def _allowed_excel(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in _ALLOWED_EXCEL


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard_redirect'))

    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        remember = request.form.get('remember') == 'on'

        conn = get_db()
        row = conn.execute(
            "SELECT * FROM usuarios WHERE username = ? AND activo = 1", (username,)
        ).fetchone()

        if row and check_password_hash(row['password_hash'], password):
            user = User(row)
            login_user(user, remember=remember)
            conn.close()
            return redirect(url_for('dashboard_redirect'))

        conn.close()
        error = 'Usuario o contraseña incorrectos.'

    return render_template('login.html', error=error)


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


@app.route('/')
@login_required
def dashboard_redirect():
    destinos = {
        'admin':      'admin_dashboard',
        'superadmin': 'admin_dashboard',
        'cio':        'cio_dashboard',
        'tecnico':    'taller',
    }
    return redirect(url_for(destinos.get(current_user.rol, 'login')))


# ---------------------------------------------------------------------------
# Admin
# ---------------------------------------------------------------------------

@app.route('/admin')
@admin_required
def admin_dashboard():
    conn = get_db()
    sync_id = _current_sync_id(conn)

    total_equipos = 0
    solicitados   = 0
    respondidos   = 0
    alerta_vencidos      = 0
    alerta_pendientes    = 0
    alerta_no_ej         = 0
    alerta_no_reportadas = 0
    sol_equipo_ids       = frozenset()

    if sync_id:
        total_equipos = conn.execute(
            "SELECT COUNT(*) AS c FROM equipos WHERE sync_id = ?", (sync_id,)
        ).fetchone()['c']

        solicitados = conn.execute(
            "SELECT COUNT(*) AS c FROM solicitudes WHERE sync_id = ?", (sync_id,)
        ).fetchone()['c']

        respondidos = conn.execute(
            """SELECT COUNT(*) AS c FROM respuestas r
               JOIN solicitudes s ON s.id = r.solicitud_id
               WHERE s.sync_id = ?""",
            (sync_id,)
        ).fetchone()['c']

    pendientes = max(0, solicitados - respondidos)

    ultimas_sync = conn.execute(
        """SELECT l.timestamp, l.detalle, u.nombre_completo
           FROM log_actividad l
           LEFT JOIN usuarios u ON u.id = l.usuario_id
           WHERE l.accion_tipo = 'sync'
           ORDER BY l.timestamp DESC
           LIMIT 10"""
    ).fetchall()

    equipos_todos = []
    categorias_admin = []
    familias_admin = []
    if sync_id:
        equipos_todos = conn.execute(
            """SELECT * FROM equipos
               WHERE sync_id = ?
               ORDER BY CAST(ind_desviacion AS INTEGER) DESC""",
            (sync_id,)
        ).fetchall()
        categorias_admin = [r['categoria'] for r in conn.execute(
            """SELECT DISTINCT categoria FROM equipos
               WHERE sync_id = ? AND categoria IS NOT NULL
               ORDER BY categoria""",
            (sync_id,)
        ).fetchall()]
        familias_admin = [r['familia'] for r in conn.execute(
            """SELECT DISTINCT familia FROM equipos
               WHERE sync_id = ? AND familia IS NOT NULL
               ORDER BY familia""",
            (sync_id,)
        ).fetchall()]

        # ── Alertas de gestión ──
        alerta_vencidos = conn.execute(
            """SELECT COUNT(DISTINCT e.vehiculo) AS c
               FROM equipos e
               WHERE e.sync_id = ?
               AND LOWER(e.estado_mp) LIKE '%vencido%'
               AND NOT EXISTS (
                   SELECT 1 FROM solicitudes s
                   WHERE s.equipo_id = e.id AND s.sync_id = e.sync_id
               )
               AND JULIANDAY('now') - JULIANDAY(e.fecha_programacion) > 7""",
            (sync_id,)
        ).fetchone()['c'] or 0

        alerta_pendientes = conn.execute(
            """SELECT COUNT(*) AS c
               FROM solicitudes s
               WHERE s.sync_id = ?
               AND s.estado = 'pendiente'
               AND JULIANDAY('now') - JULIANDAY(s.fecha_solicitud) > 3""",
            (sync_id,)
        ).fetchone()['c'] or 0

        alerta_no_ej = conn.execute(
            """SELECT COUNT(DISTINCT e.vehiculo) AS c
               FROM solicitudes s
               JOIN equipos e ON e.id = s.equipo_id
               JOIN respuestas r ON r.solicitud_id = s.id
               WHERE s.sync_id = ?
               AND r.accion = 'no_ejecutado'""",
            (sync_id,)
        ).fetchone()['c'] or 0

        sol_equipo_ids = frozenset(
            r['equipo_id'] for r in conn.execute(
                "SELECT DISTINCT equipo_id FROM solicitudes WHERE sync_id = ?",
                (sync_id,)
            ).fetchall()
        )

        alerta_no_reportadas = conn.execute(
            "SELECT COUNT(*) AS c FROM ejecuciones_no_reportadas WHERE estado = 'pendiente'"
        ).fetchone()['c'] or 0

    conn.close()

    return render_template('admin/dashboard.html',
        total_equipos=total_equipos,
        solicitados=solicitados,
        respondidos=respondidos,
        pendientes=pendientes,
        ultimas_sync=ultimas_sync,
        current_sync_id=sync_id,
        equipos_todos=equipos_todos,
        categorias_admin=categorias_admin,
        familias_admin=familias_admin,
        alerta_vencidos=alerta_vencidos,
        alerta_pendientes=alerta_pendientes,
        alerta_no_ej=alerta_no_ej,
        alerta_no_reportadas=alerta_no_reportadas,
        sol_equipo_ids=sol_equipo_ids,
    )


@app.route('/admin/sync', methods=['GET', 'POST'])
@admin_required
def admin_sync():
    if request.method == 'GET':
        return render_template('admin/sync.html')

    file_prog = request.files.get('file_programacion')
    file_filt = request.files.get('file_filtros')

    has_prog = file_prog and file_prog.filename
    has_filt = file_filt and file_filt.filename

    if not has_prog and not has_filt:
        flash('Debes subir al menos un archivo.', 'error')
        return render_template('admin/sync.html')

    os.makedirs(config.UPLOAD_FOLDER, exist_ok=True)

    if has_prog:
        if not _allowed_excel(file_prog.filename):
            flash('Programación MP: formato no válido. Solo se aceptan .xlsx o .xls', 'error')
        else:
            save_path = os.path.join(config.UPLOAD_FOLDER, 'programacion_mp.xlsx')
            file_prog.save(save_path)
            try:
                res = sync_data.sync_programacion(save_path)
                msg = (f'Programación MP sincronizada: {res["nuevos"]} nuevos, '
                       f'{res["actualizados"]} actualizados — {res["total"]} equipos totales '
                       f'(ciclo {res["sync_id"]})')
                flash(msg, 'success')
                if res.get('no_reportadas', 0) > 0:
                    flash(
                        f'{res["no_reportadas"]} equipo(s) detectado(s) como ejecutados '
                        f'sin pasar por el sistema. Revisa "No Reportadas".',
                        'warning'
                    )
            except Exception as exc:
                flash(f'Error en programación MP: {exc}', 'error')
            else:
                try:
                    os.remove(save_path)
                except Exception as exc:
                    app.logger.error('No se pudo eliminar programacion_mp.xlsx: %s', exc)

    if has_filt:
        if not _allowed_excel(file_filt.filename):
            flash('Maestro Filtración: formato no válido. Solo se aceptan .xlsx o .xls', 'error')
        else:
            save_path = os.path.join(config.UPLOAD_FOLDER, 'maestro_filtracion.xlsx')
            file_filt.save(save_path)
            try:
                res = sync_data.sync_filtros(save_path)
                msg = (f'Maestro Filtración sincronizado: {res["total_registros"]} registros, '
                       f'{res["equipos_unicos"]} equipos únicos')
                flash(msg, 'success')
            except Exception as exc:
                flash(f'Error en maestro filtración: {exc}', 'error')
            else:
                try:
                    os.remove(save_path)
                except Exception as exc:
                    app.logger.error('No se pudo eliminar maestro_filtracion.xlsx: %s', exc)

    return redirect(url_for('admin_sync'))


# ---------------------------------------------------------------------------
# Admin — historial, export, usuarios
# ---------------------------------------------------------------------------

@app.route('/admin/historial')
@admin_required
def admin_historial():
    page        = max(1, request.args.get('page', 1, type=int))
    per_page    = 50
    fecha_desde = request.args.get('fecha_desde', '').strip()
    fecha_hasta = request.args.get('fecha_hasta', '').strip()
    accion_fil  = request.args.get('accion', '').strip()
    familia_fil = request.args.get('familia', '').strip()
    vehiculo_fil = request.args.get('vehiculo', '').strip()

    where_parts, params = [], []

    if fecha_desde:
        where_parts.append("s.fecha_solicitud >= ?")
        params.append(fecha_desde)
    if fecha_hasta:
        where_parts.append("s.fecha_solicitud <= ?")
        params.append(fecha_hasta + 'T23:59:59')
    if accion_fil == 'pendiente':
        where_parts.append("s.estado = 'pendiente'")
    elif accion_fil in ('ejecutado', 'no_ejecutado'):
        where_parts.append("r.accion = ?")
        params.append(accion_fil)
    if familia_fil:
        where_parts.append("e.familia = ?")
        params.append(familia_fil)
    if vehiculo_fil:
        where_parts.append("UPPER(e.vehiculo) LIKE ?")
        params.append(f'%{vehiculo_fil.upper()}%')

    where_sql = ('WHERE ' + ' AND '.join(where_parts)) if where_parts else ''

    base_from = f"""
        FROM solicitudes s
        JOIN equipos e   ON e.id = s.equipo_id
        JOIN usuarios u  ON u.id = s.solicitado_por
        LEFT JOIN respuestas r       ON r.solicitud_id = s.id
        LEFT JOIN catalogo_motivos m ON m.id = r.motivo_id
        {where_sql}
    """

    conn = get_db()
    total  = conn.execute(f"SELECT COUNT(*) AS c {base_from}", params).fetchone()['c']
    offset = (page - 1) * per_page

    filas = conn.execute(
        f"""SELECT s.id, s.fecha_solicitud, s.estado,
                   e.vehiculo, e.familia, e.rutina, e.desviacion, e.ind_desviacion, e.estado_mp,
                   u.nombre_completo AS solicitado_por,
                   r.accion, r.timestamp AS fecha_respuesta,
                   m.descripcion AS motivo, r.comentario_libre
            {base_from}
            ORDER BY s.fecha_solicitud DESC
            LIMIT ? OFFSET ?""",
        params + [per_page, offset]
    ).fetchall()

    familias = [row['familia'] for row in conn.execute(
        "SELECT DISTINCT familia FROM equipos WHERE familia IS NOT NULL ORDER BY familia"
    ).fetchall()]
    conn.close()

    total_pages = max(1, (total + per_page - 1) // per_page)

    return render_template('admin/historial.html',
        filas=filas,
        total=total,
        page=page,
        per_page=per_page,
        total_pages=total_pages,
        familias=familias,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        accion_fil=accion_fil,
        familia_fil=familia_fil,
        vehiculo_fil=vehiculo_fil,
    )


@app.route('/admin/export')
@admin_required
def admin_export():
    import io
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    conn = get_db()
    filas = conn.execute(
        """SELECT s.fecha_solicitud, e.vehiculo, e.familia, e.categoria, e.rutina,
                  e.desviacion, e.ind_desviacion, e.estado_mp,
                  u.nombre_completo AS solicitado_por,
                  r.timestamp AS fecha_respuesta, r.accion,
                  m.descripcion AS motivo, r.comentario_libre
           FROM solicitudes s
           JOIN equipos e  ON e.id = s.equipo_id
           JOIN usuarios u ON u.id = s.solicitado_por
           LEFT JOIN respuestas r       ON r.solicitud_id = s.id
           LEFT JOIN catalogo_motivos m ON m.id = r.motivo_id
           ORDER BY s.fecha_solicitud DESC"""
    ).fetchall()

    # ── Datos para Resumen Ejecutivo ──
    stats = conn.execute(
        """SELECT
               COUNT(s.id) AS total_sol,
               SUM(CASE WHEN r.accion = 'ejecutado'    THEN 1 ELSE 0 END) AS ejecutados,
               SUM(CASE WHEN r.accion = 'no_ejecutado' THEN 1 ELSE 0 END) AS no_ejecutados
           FROM solicitudes s
           LEFT JOIN respuestas r ON r.solicitud_id = s.id"""
    ).fetchone()
    total_sol  = stats['total_sol']      or 0
    ejecutados = stats['ejecutados']     or 0
    no_ej      = stats['no_ejecutados']  or 0
    pct_ej     = round(ejecutados / total_sol * 100) if total_sol else 0

    familias_data = conn.execute(
        """SELECT e.familia,
                  COUNT(s.id) AS solicitados,
                  SUM(CASE WHEN r.accion = 'ejecutado'    THEN 1 ELSE 0 END) AS ejecutados,
                  SUM(CASE WHEN r.accion = 'no_ejecutado' THEN 1 ELSE 0 END) AS no_ejecutados
           FROM solicitudes s
           JOIN equipos e ON e.id = s.equipo_id
           LEFT JOIN respuestas r ON r.solicitud_id = s.id
           WHERE e.familia IS NOT NULL
           GROUP BY e.familia
           ORDER BY e.familia"""
    ).fetchall()

    top_motivos = conn.execute(
        """SELECT CASE WHEN r.motivo_id IS NOT NULL
                       THEN COALESCE(m.descripcion, 'Desconocido')
                       ELSE 'Comentario libre'
                  END AS motivo,
                  COUNT(*) AS cantidad
           FROM respuestas r
           LEFT JOIN catalogo_motivos m ON m.id = r.motivo_id
           WHERE r.accion = 'no_ejecutado'
           GROUP BY r.motivo_id
           ORDER BY cantidad DESC
           LIMIT 10"""
    ).fetchall()
    conn.close()

    # ── Estilos comunes ──
    thin           = Side(style='thin', color='CCCCCC')
    brd            = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_azul      = PatternFill('solid', fgColor='002D6E')
    fill_verde     = PatternFill('solid', fgColor='80AE3F')
    fill_cielo     = PatternFill('solid', fgColor='1E88E5')
    fill_ambar     = PatternFill('solid', fgColor='E67E22')
    fill_gris      = PatternFill('solid', fgColor='F0F2F5')
    fill_seccion   = PatternFill('solid', fgColor='E8EDF5')
    font_hdr       = Font(bold=True, color='FFFFFF', size=10)
    font_seccion   = Font(bold=True, color='002D6E', size=11)
    font_body      = Font(size=10)
    center         = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_al        = Alignment(horizontal='left',   vertical='center')

    wb = openpyxl.Workbook()

    # ════════════════════════════════════════
    # HOJA 1 — Resumen Ejecutivo
    # ════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Resumen Ejecutivo'

    # Encabezado corporativo
    ws1.merge_cells('A1:F1')
    c = ws1['A1']
    c.value     = 'GET — Talma Servicios Aeroportuarios'
    c.fill      = fill_azul
    c.font      = Font(bold=True, color='FFFFFF', size=16)
    c.alignment = center
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells('A2:F2')
    c2 = ws1['A2']
    c2.value     = 'Reporte Seguimiento Mantenimiento Preventivo GSE — BOG'
    c2.fill      = fill_verde
    c2.font      = Font(bold=True, color='FFFFFF', size=11)
    c2.alignment = center
    ws1.row_dimensions[2].height = 22

    ws1.merge_cells('A3:F3')
    c3 = ws1['A3']
    fecha_gen    = datetime.now(TZ_COL).strftime('%d/%m/%Y %H:%M')
    c3.value     = f'Generado: {fecha_gen} (hora Colombia)'
    c3.fill      = fill_gris
    c3.font      = Font(italic=True, color='6B7280', size=9)
    c3.alignment = center
    ws1.row_dimensions[3].height = 16
    ws1.row_dimensions[4].height = 8  # separador

    # Sección KPIs
    ws1.merge_cells('A5:F5')
    cs = ws1['A5']
    cs.value     = '   INDICADORES GENERALES'
    cs.fill      = fill_seccion
    cs.font      = font_seccion
    cs.alignment = left_al
    ws1.row_dimensions[5].height = 20

    kpi_data = [
        ('Total Solicitados', total_sol,     fill_cielo),
        ('Ejecutados',        ejecutados,    fill_verde),
        ('No Ejecutados',     no_ej,         fill_ambar),
        ('% Ejecución',       f'{pct_ej}%',  fill_azul),
    ]
    for ci, (lbl, val, fill) in enumerate(kpi_data, 1):
        hc = ws1.cell(row=6, column=ci, value=lbl)
        hc.fill = fill; hc.font = font_hdr; hc.alignment = center
        ws1.row_dimensions[6].height = 18
        vc = ws1.cell(row=7, column=ci, value=val)
        vc.fill      = PatternFill('solid', fgColor='FFFFFF')
        vc.font      = Font(bold=True, color='002D6E', size=24)
        vc.alignment = center
        vc.border    = brd
        ws1.row_dimensions[7].height = 40

    ws1.row_dimensions[8].height = 8

    # Sección Familias
    ws1.merge_cells('A9:F9')
    cf = ws1['A9']
    cf.value     = '   CUMPLIMIENTO POR FAMILIA'
    cf.fill      = fill_seccion
    cf.font      = font_seccion
    cf.alignment = left_al
    ws1.row_dimensions[9].height = 20

    fam_hdrs = ['Familia', 'Solicitados', 'Ejecutados', 'No Ejecutados', '% Ejecución']
    for ci, h in enumerate(fam_hdrs, 1):
        cell = ws1.cell(row=10, column=ci, value=h)
        cell.fill = fill_azul; cell.font = font_hdr
        cell.alignment = center; cell.border = brd
    ws1.row_dimensions[10].height = 20

    for ri, fd in enumerate(familias_data, 11):
        sf = fd['solicitados'] or 0
        ef = fd['ejecutados']  or 0
        nf = fd['no_ejecutados'] or 0
        pf = f"{round(ef/sf*100)}%" if sf else '—'
        for ci, v in enumerate([fd['familia'] or '—', sf, ef, nf, pf], 1):
            cell = ws1.cell(row=ri, column=ci, value=v)
            cell.alignment = center; cell.border = brd; cell.font = font_body
            if ri % 2 == 0: cell.fill = fill_gris
        ws1.row_dimensions[ri].height = 16

    last_fam = 10 + len(familias_data) + 1
    ws1.row_dimensions[last_fam].height = 8

    # Sección Top Motivos
    mr = last_fam + 1
    ws1.merge_cells(f'A{mr}:F{mr}')
    cm = ws1[f'A{mr}']
    cm.value     = '   TOP MOTIVOS DE NO EJECUCIÓN'
    cm.fill      = fill_seccion
    cm.font      = font_seccion
    cm.alignment = left_al
    ws1.row_dimensions[mr].height = 20

    mhr = mr + 1
    for ci, h in enumerate(['Motivo', 'Cantidad'], 1):
        cell = ws1.cell(row=mhr, column=ci, value=h)
        cell.fill = fill_ambar; cell.font = font_hdr
        cell.alignment = center; cell.border = brd
    ws1.row_dimensions[mhr].height = 20

    for ri, m in enumerate(top_motivos, mhr + 1):
        for ci, v in enumerate([m['motivo'] or 'Sin motivo', m['cantidad']], 1):
            cell = ws1.cell(row=ri, column=ci, value=v)
            cell.alignment = left_al if ci == 1 else center
            cell.border = brd; cell.font = font_body
            if ri % 2 == 0: cell.fill = fill_gris
        ws1.row_dimensions[ri].height = 16

    for col_letter, w in [('A',36),('B',16),('C',16),('D',18),('E',14),('F',14)]:
        ws1.column_dimensions[col_letter].width = w

    # ════════════════════════════════════════
    # HOJA 2 — Detalle
    # ════════════════════════════════════════
    ws2 = wb.create_sheet(title='Detalle')

    det_headers = [
        'Fecha Solicitud', 'Vehículo', 'Familia', 'Categoría', 'Rutina',
        'Desviación', 'Ind. Desviación', 'Estado MP',
        'Solicitado por', 'Fecha Respuesta', 'Acción', 'Motivo', 'Comentario',
    ]
    ws2.row_dimensions[1].height = 28
    for ci, h in enumerate(det_headers, 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.fill = fill_azul; cell.font = font_hdr
        cell.alignment = center; cell.border = brd

    body_align = Alignment(vertical='center')
    for ri, fila in enumerate(filas, 2):
        values = [
            fila['fecha_solicitud'], fila['vehiculo'], fila['familia'],
            fila['categoria'], fila['rutina'], fila['desviacion'],
            fila['ind_desviacion'], fila['estado_mp'], fila['solicitado_por'],
            fila['fecha_respuesta'], fila['accion'] or 'pendiente',
            fila['motivo'], fila['comentario_libre'],
        ]
        for ci, val in enumerate(values, 1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.border = brd; cell.alignment = body_align; cell.font = font_body
            if ri % 2 == 0: cell.fill = fill_gris

    ws2.auto_filter.ref = ws2.dimensions
    ws2.freeze_panes    = 'A2'

    for col in ws2.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0 for c in col),
            default=10
        )
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 55)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"reporte_mp_{datetime.now(TZ_COL).strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


@app.route('/admin/usuarios', methods=['GET', 'POST'])
@admin_required
def admin_usuarios():
    is_superadmin = current_user.rol == 'superadmin'
    valid_roles_create = ('admin', 'cio', 'tecnico', 'superadmin') if is_superadmin else ('admin', 'cio', 'tecnico')

    if request.method == 'POST':
        action = request.form.get('action', '')

        if action == 'eliminar':
            uid = request.form.get('user_id', type=int)
            if uid == current_user.id:
                flash('No puedes eliminar tu propia cuenta.', 'error')
                return redirect(url_for('admin_usuarios'))
            if uid:
                conn = get_db()
                try:
                    row = conn.execute(
                        "SELECT username, rol FROM usuarios WHERE id = ?", (uid,)
                    ).fetchone()
                    if not row:
                        flash('Usuario no encontrado.', 'error')
                    elif not is_superadmin and row['rol'] == 'superadmin':
                        flash('No tienes permiso para eliminar cuentas superadmin.', 'error')
                    else:
                        n_sol = conn.execute(
                            "SELECT COUNT(*) AS c FROM solicitudes WHERE solicitado_por = ?", (uid,)
                        ).fetchone()['c']
                        if n_sol > 0:
                            conn.close()
                            flash(f'No se puede eliminar: tiene {n_sol} solicitud(es) registrada(s). '
                                  'Desactívelo en su lugar.', 'error')
                            return redirect(url_for('admin_usuarios'))
                        conn.execute("DELETE FROM usuarios WHERE id = ?", (uid,))
                        conn.commit()
                        flash(f'Usuario "{row["username"]}" eliminado permanentemente.', 'success')
                finally:
                    conn.close()
            return redirect(url_for('admin_usuarios'))

        conn = get_db()
        try:
            if action == 'crear':
                username = request.form.get('username', '').strip().lower()
                password = request.form.get('password', '').strip()
                nombre   = request.form.get('nombre_completo', '').strip()
                rol      = request.form.get('rol', '').strip()

                errores = []
                if not username:
                    errores.append('El username es obligatorio.')
                if not password or len(password) < 6:
                    errores.append('La contraseña debe tener mínimo 6 caracteres.')
                if not nombre:
                    errores.append('El nombre completo es obligatorio.')
                if rol not in valid_roles_create:
                    errores.append('Rol inválido.')

                if not errores:
                    dup = conn.execute(
                        "SELECT id FROM usuarios WHERE username = ?", (username,)
                    ).fetchone()
                    if dup:
                        errores.append(f'El username "{username}" ya existe.')

                if errores:
                    for e in errores:
                        flash(e, 'error')
                else:
                    conn.execute(
                        """INSERT INTO usuarios
                               (username, password_hash, nombre_completo, rol, activo, created_at)
                           VALUES (?, ?, ?, ?, 1, ?)""",
                        (username, generate_password_hash(password), nombre, rol,
                         datetime.now(TZ_COL).isoformat())
                    )
                    conn.commit()
                    flash(f'Usuario "{username}" creado exitosamente.', 'success')

            elif action == 'toggle':
                uid = request.form.get('user_id', type=int)
                if uid == current_user.id:
                    flash('No puedes modificar tu propia cuenta.', 'error')
                elif uid:
                    row = conn.execute(
                        "SELECT activo, username, rol FROM usuarios WHERE id = ?", (uid,)
                    ).fetchone()
                    if row:
                        if not is_superadmin and row['rol'] == 'superadmin':
                            flash('No tienes permiso para modificar cuentas superadmin.', 'error')
                        else:
                            nuevo = 0 if row['activo'] else 1
                            conn.execute(
                                "UPDATE usuarios SET activo = ? WHERE id = ?", (nuevo, uid)
                            )
                            verb = 'activado' if nuevo else 'desactivado'
                            conn.commit()
                            flash(f'Usuario "{row["username"]}" {verb}.', 'success')
        finally:
            conn.close()
        return redirect(url_for('admin_usuarios'))

    conn = get_db()
    try:
        if is_superadmin:
            usuarios = conn.execute(
                "SELECT * FROM usuarios ORDER BY created_at DESC"
            ).fetchall()
        else:
            usuarios = conn.execute(
                "SELECT * FROM usuarios WHERE rol != 'superadmin' ORDER BY created_at DESC"
            ).fetchall()
    finally:
        conn.close()

    return render_template('admin/usuarios.html',
        usuarios=usuarios,
        valid_roles_create=valid_roles_create,
    )


# ---------------------------------------------------------------------------
# CIO
# ---------------------------------------------------------------------------

@app.route('/cio')
@cio_required
def cio_dashboard():
    conn = get_db()
    sync_id = _current_sync_id(conn)

    equipos_grouped = []
    familias, estados, categorias, estados_vehiculo = [], [], [], []
    total = vencidos = proximos = ya_solicitados = ejecutados_count = 0
    ultima_actualizacion = None

    if sync_id:
        rows_sol = conn.execute(
            """SELECT s.id, s.equipo_id, s.estado, e.vehiculo,
                      r.accion, m.descripcion AS motivo_desc, r.comentario_libre
               FROM solicitudes s
               JOIN equipos e ON e.id = s.equipo_id
               LEFT JOIN respuestas r ON r.solicitud_id = s.id
               LEFT JOIN catalogo_motivos m ON m.id = r.motivo_id
               WHERE s.sync_id = ?""",
            (sync_id,)
        ).fetchall()

        sol_by_equipo = {}
        for row in rows_sol:
            eid = row['equipo_id']
            existing = sol_by_equipo.get(eid)
            entry = {
                'sol_id':           row['id'],
                'estado':           row['estado'],
                'accion':           row['accion'],
                'motivo_desc':      row['motivo_desc'],
                'comentario_libre': row['comentario_libre'],
            }
            if existing is None:
                sol_by_equipo[eid] = entry
            elif existing['estado'] == 'respondido' and row['estado'] == 'pendiente':
                sol_by_equipo[eid] = entry

        all_equipos = conn.execute(
            """SELECT * FROM equipos
               WHERE sync_id = ?
               AND CAST(ind_desviacion AS INTEGER) >= -10
               ORDER BY CAST(ind_desviacion AS INTEGER) DESC""",
            (sync_id,)
        ).fetchall()

        vehicles = {}
        for equipo in all_equipos:
            sol = sol_by_equipo.get(equipo['id'])
            if sol and sol['accion'] == 'ejecutado' and sol['estado'] == 'respondido':
                continue

            v = equipo['vehiculo']
            if v not in vehicles:
                vehicles[v] = {
                    'vehiculo':           v,
                    'familia':            equipo['familia'],
                    'categoria':          equipo['categoria'],
                    'estado_vehiculo':    equipo['estado_vehiculo'],
                    'rutinas':            [],
                    'ind_desviacion':     equipo['ind_desviacion'],
                    'desviacion':         equipo['desviacion'],
                    'estado_mp':          equipo['estado_mp'],
                    'equipo_ids':         [],
                    'sol_ids_pendientes': [],
                    'motivos_no_ej':      [],
                    '_has_pendiente':     False,
                    '_has_no_ej':         False,
                    'fecha_programacion': equipo['fecha_programacion'],
                }

            vd = vehicles[v]
            vd['equipo_ids'].append(equipo['id'])
            if equipo['rutina']:
                vd['rutinas'].append(equipo['rutina'])

            em   = (equipo['estado_mp'] or '').lower()
            curr = (vd['estado_mp'] or '').lower()
            if 'vencido' in em and 'vencido' not in curr:
                vd['estado_mp'] = equipo['estado_mp']

            if sol:
                if sol['estado'] == 'pendiente':
                    vd['sol_ids_pendientes'].append(sol['sol_id'])
                    vd['_has_pendiente'] = True
                elif sol['accion'] == 'no_ejecutado':
                    mot = sol['motivo_desc'] or sol['comentario_libre'] or 'Sin motivo'
                    if mot not in vd['motivos_no_ej']:
                        vd['motivos_no_ej'].append(mot)
                    vd['_has_no_ej'] = True

        # Calcular días sin gestionar para vehículos vencidos sin solicitud
        _today = datetime.now(TZ_COL).date()
        for _vd in vehicles.values():
            _vd['dias_sin_gestionar'] = 0
            if _vd['_has_pendiente'] or _vd['_has_no_ej']:
                continue
            if 'vencido' not in (_vd['estado_mp'] or '').lower():
                continue
            fp = _vd.get('fecha_programacion')
            if not fp:
                continue
            try:
                d = datetime.strptime(str(fp)[:10], '%Y-%m-%d').date()
                diff = (_today - d).days
                if diff > 0:
                    _vd['dias_sin_gestionar'] = diff
            except (ValueError, TypeError):
                pass

        for vd in vehicles.values():
            if vd['_has_pendiente']:
                vd['sol_state'] = 'pendiente'
            elif vd['_has_no_ej']:
                vd['sol_state'] = 'no_ejecutado'
            else:
                vd['sol_state'] = None
            vd['rutinas_str']      = ', '.join(vd['rutinas'])
            vd['motivo_no_ej_str'] = ' / '.join(vd['motivos_no_ej'])
            del vd['_has_pendiente'], vd['_has_no_ej'], vd['fecha_programacion']

        def _sk(vd):
            try:
                return -(int(str(vd['ind_desviacion']).replace('%', '').strip())
                         if vd['ind_desviacion'] is not None else 0)
            except (ValueError, TypeError):
                return 0

        equipos_grouped = sorted(vehicles.values(), key=_sk)

        _seen = {'fam': set(), 'est': set(), 'cat': set(), 'eveh': set()}
        for vd in equipos_grouped:
            if vd['familia'] and vd['familia'] not in _seen['fam']:
                familias.append(vd['familia']); _seen['fam'].add(vd['familia'])
            if vd['estado_mp'] and vd['estado_mp'] not in _seen['est']:
                estados.append(vd['estado_mp']); _seen['est'].add(vd['estado_mp'])
            if vd['categoria'] and vd['categoria'] not in _seen['cat']:
                categorias.append(vd['categoria']); _seen['cat'].add(vd['categoria'])
            if vd['estado_vehiculo'] and vd['estado_vehiculo'] not in _seen['eveh']:
                estados_vehiculo.append(vd['estado_vehiculo']); _seen['eveh'].add(vd['estado_vehiculo'])
        familias.sort(); estados.sort(); categorias.sort(); estados_vehiculo.sort()

        total          = len(equipos_grouped)
        vencidos       = sum(1 for vd in equipos_grouped
                             if vd['estado_mp'] and 'vencido' in vd['estado_mp'].lower())
        proximos       = sum(1 for vd in equipos_grouped
                             if vd['estado_mp'] and
                             ('próximo' in vd['estado_mp'].lower() or 'proximo' in vd['estado_mp'].lower()))
        ya_solicitados = sum(1 for vd in equipos_grouped if vd['sol_state'] is not None)

        veh_sol_acciones = {}
        for row in rows_sol:
            veh = row['vehiculo']
            if veh not in veh_sol_acciones:
                veh_sol_acciones[veh] = []
            if row['accion']:
                veh_sol_acciones[veh].append(row['accion'])

        ejecutados_count = sum(
            1 for veh, acciones in veh_sol_acciones.items()
            if acciones and all(a == 'ejecutado' for a in acciones)
            and veh not in vehicles
        )

        row_ult = conn.execute(
            "SELECT fecha_programacion FROM equipos ORDER BY sync_timestamp DESC LIMIT 1"
        ).fetchone()
        ultima_actualizacion = _format_fecha_actualizacion(row_ult['fecha_programacion']) if row_ult else None

    conn.close()

    return render_template('cio/dashboard.html',
        equipos=equipos_grouped,
        familias=familias,
        estados=estados,
        categorias=categorias,
        estados_vehiculo=estados_vehiculo,
        total=total,
        vencidos=vencidos,
        proximos=proximos,
        ya_solicitados=ya_solicitados,
        ejecutados_count=ejecutados_count,
        current_sync_id=sync_id,
        ultima_actualizacion=ultima_actualizacion,
    )


@app.route('/cio/solicitar', methods=['POST'])
@cio_required
def cio_solicitar():
    vehiculos = request.form.getlist('vehiculos')
    if not vehiculos:
        flash('No seleccionaste ningún vehículo.', 'error')
        return redirect(url_for('cio_dashboard'))

    conn = get_db()
    sync_id = _current_sync_id(conn)
    ahora   = datetime.now(TZ_COL).isoformat()
    registrados = 0

    for vehiculo in vehiculos:
        vehiculo = vehiculo.strip().upper()
        equipo_rows = conn.execute(
            """SELECT id FROM equipos
               WHERE UPPER(vehiculo) = ? AND sync_id = ?
               AND CAST(ind_desviacion AS INTEGER) >= -10""",
            (vehiculo, sync_id)
        ).fetchall()

        for row in equipo_rows:
            eid = row['id']
            existing_pendiente = conn.execute(
                """SELECT id FROM solicitudes
                   WHERE equipo_id = ? AND sync_id = ? AND estado = 'pendiente'""",
                (eid, sync_id)
            ).fetchone()
            if not existing_pendiente:
                conn.execute(
                    """INSERT INTO solicitudes
                           (equipo_id, solicitado_por, fecha_solicitud, sync_id, estado)
                       VALUES (?, ?, ?, ?, 'pendiente')""",
                    (eid, current_user.id, ahora, sync_id)
                )
                registrados += 1

    if registrados:
        _log_actividad(conn, current_user.id, 'solicitud',
                       f'Solicitó {registrados} rutina(s): {", ".join(vehiculos)}')
    conn.commit()
    conn.close()

    if registrados:
        flash(f'{len(vehiculos)} vehículo(s) solicitados — {registrados} rutina(s) registradas. '
              'Notifica a Operaciones para coordinar la entrega.', 'success')
    else:
        flash('Los vehículos seleccionados ya tenían solicitudes pendientes en este ciclo.', 'warning')

    return redirect(url_for('cio_dashboard'))


@app.route('/cio/motivos')
@cio_required
def cio_motivos():
    conn = get_db()
    motivos = conn.execute(
        "SELECT id, descripcion FROM catalogo_motivos WHERE activo = 1 ORDER BY orden"
    ).fetchall()
    conn.close()
    return jsonify([{'id': m['id'], 'descripcion': m['descripcion']} for m in motivos])


@app.route('/cio/responder', methods=['POST'])
@cio_required
def cio_responder():
    data        = request.get_json(force=True) or {}
    vehiculo    = (data.get('vehiculo') or '').strip().upper()
    accion      = data.get('accion')
    motivo_id   = data.get('motivo_id')
    comentario  = (data.get('comentario_libre') or '').strip() or None

    if not vehiculo or accion not in ('ejecutado', 'no_ejecutado'):
        return jsonify({'success': False, 'error': 'Datos inválidos.'}), 400

    if accion == 'no_ejecutado' and not motivo_id:
        return jsonify({'success': False, 'error': 'Motivo obligatorio para "No ejecutado".'}), 400

    conn = get_db()
    sync_id = _current_sync_id(conn)

    solicitudes = conn.execute(
        """SELECT s.id FROM solicitudes s
           JOIN equipos e ON e.id = s.equipo_id
           WHERE UPPER(e.vehiculo) = ? AND s.sync_id = ? AND s.estado = 'pendiente'""",
        (vehiculo, sync_id)
    ).fetchall()

    if not solicitudes:
        conn.close()
        return jsonify({'success': False, 'error': 'No hay solicitudes pendientes para este vehículo.'}), 404

    motivo_desc = None
    if motivo_id:
        m = conn.execute(
            "SELECT descripcion FROM catalogo_motivos WHERE id = ?", (motivo_id,)
        ).fetchone()
        if m:
            motivo_desc = m['descripcion']

    ahora = datetime.now(TZ_COL).isoformat()
    for sol in solicitudes:
        conn.execute(
            """INSERT INTO respuestas
                   (solicitud_id, respondido_por, accion, motivo_id, comentario_libre, timestamp, ip_address)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (sol['id'], current_user.id, accion, motivo_id or None, comentario, ahora, request.remote_addr)
        )
        conn.execute("UPDATE solicitudes SET estado = 'respondido' WHERE id = ?", (sol['id'],))

    conn.commit()
    conn.close()
    return jsonify({'success': True, 'accion': accion, 'motivo_desc': motivo_desc})


@app.route('/cio/mis-solicitudes')
@cio_required
def cio_mis_solicitudes():
    conn = get_db()
    sync_id = _current_sync_id(conn)

    solicitudes = conn.execute(
        """SELECT s.id, s.fecha_solicitud, s.estado, e.vehiculo, e.familia,
                  e.rutina, e.estado_mp,
                  r.accion, r.timestamp AS resp_timestamp,
                  m.descripcion AS motivo, r.comentario_libre
           FROM solicitudes s
           JOIN equipos e ON e.id = s.equipo_id
           LEFT JOIN respuestas r ON r.solicitud_id = s.id
           LEFT JOIN catalogo_motivos m ON m.id = r.motivo_id
           WHERE s.solicitado_por = ?
           ORDER BY s.fecha_solicitud DESC
           LIMIT 100""",
        (current_user.id,)
    ).fetchall()

    conn.close()

    return render_template('cio/mis_solicitudes.html',
        solicitudes=solicitudes,
        current_sync_id=sync_id,
    )


# ---------------------------------------------------------------------------
# Equipo detalle (accesible por ambos roles)
# ---------------------------------------------------------------------------

@app.route('/equipo/<vehiculo>')
@login_required
def equipo_detalle(vehiculo):
    vehiculo = vehiculo.upper().strip()
    from_page = request.args.get('origen', '')
    conn = get_db()
    sync_id = _current_sync_id(conn)

    rutinas = conn.execute(
        """SELECT * FROM equipos
           WHERE UPPER(vehiculo) = ? AND sync_id = ?
           ORDER BY ind_desviacion DESC""",
        (vehiculo, sync_id)
    ).fetchall() if sync_id else []

    filtros = conn.execute(
        """SELECT * FROM filtros_equipo
           WHERE UPPER(equipo) = ?
           ORDER BY tipo_filtro, nombre_articulo""",
        (vehiculo,)
    ).fetchall()

    sugerencias = conn.execute(
        """SELECT sf.id, sf.vehiculo, sf.descripcion, sf.estado, sf.timestamp, sf.respuesta_admin,
                  fe.nombre_articulo, fe.tipo_filtro,
                  u.nombre_completo, u.rol
           FROM sugerencias_filtros sf
           JOIN usuarios u ON u.id = sf.usuario_id
           LEFT JOIN filtros_equipo fe ON fe.id = sf.filtro_id
           WHERE UPPER(sf.vehiculo) = ?
           ORDER BY sf.timestamp DESC
           LIMIT 20""",
        (vehiculo,)
    ).fetchall()

    historial = conn.execute(
        """SELECT s.fecha_solicitud, s.estado,
                  u.nombre_completo AS solicitado_por,
                  r.accion, r.timestamp AS resp_timestamp,
                  m.descripcion AS motivo, r.comentario_libre
           FROM solicitudes s
           JOIN equipos e ON e.id = s.equipo_id
           JOIN usuarios u ON u.id = s.solicitado_por
           LEFT JOIN respuestas r ON r.solicitud_id = s.id
           LEFT JOIN catalogo_motivos m ON m.id = r.motivo_id
           WHERE UPPER(e.vehiculo) = ?
           ORDER BY s.fecha_solicitud DESC
           LIMIT 20""",
        (vehiculo,)
    ).fetchall()

    conn.close()

    if not rutinas and not filtros:
        flash(f'No se encontraron datos para el vehículo {vehiculo}.', 'warning')
        return redirect(url_for('dashboard_redirect'))

    equipo_base = rutinas[0] if rutinas else None

    hist_stats = {
        'total':         len(historial),
        'ejecutados':    sum(1 for h in historial if h['accion'] == 'ejecutado'),
        'no_ejecutados': sum(1 for h in historial if h['accion'] == 'no_ejecutado'),
        'pendientes':    sum(1 for h in historial if h['estado'] == 'pendiente'),
    }
    hist_stats['pct_ejecucion'] = (
        round(hist_stats['ejecutados'] / hist_stats['total'] * 100)
        if hist_stats['total'] else 0
    )

    return render_template('equipo_detalle.html',
        vehiculo=vehiculo,
        equipo=equipo_base,
        rutinas=rutinas,
        filtros=filtros,
        historial=historial,
        sugerencias=sugerencias,
        from_page=from_page,
        hist_stats=hist_stats,
    )


# ---------------------------------------------------------------------------
# Sugerencias de filtración
# ---------------------------------------------------------------------------

@app.route('/equipo/<vehiculo>/sugerencia', methods=['POST'])
@login_required
def equipo_sugerencia(vehiculo):
    vehiculo = vehiculo.upper().strip()
    filtro_id = request.form.get('filtro_id') or None
    if filtro_id:
        try:
            filtro_id = int(filtro_id)
        except (ValueError, TypeError):
            filtro_id = None
    descripcion = request.form.get('descripcion', '').strip()

    if not descripcion:
        flash('La descripción de la sugerencia es obligatoria.', 'error')
        return redirect(url_for('equipo_detalle', vehiculo=vehiculo))

    conn = get_db()
    conn.execute(
        """INSERT INTO sugerencias_filtros
               (vehiculo, filtro_id, usuario_id, descripcion, estado, timestamp)
           VALUES (?, ?, ?, ?, 'pendiente', ?)""",
        (vehiculo, filtro_id, current_user.id, descripcion, datetime.now(TZ_COL).isoformat())
    )
    conn.commit()
    conn.close()
    flash('Sugerencia enviada correctamente.', 'success')
    return redirect(url_for('equipo_detalle', vehiculo=vehiculo))


@app.route('/admin/sugerencias')
@admin_required
def admin_sugerencias():
    estado_fil   = request.args.get('estado', '').strip()
    vehiculo_fil = request.args.get('vehiculo', '').strip()

    where_parts, params = [], []
    if estado_fil:
        where_parts.append("sf.estado = ?")
        params.append(estado_fil)
    if vehiculo_fil:
        where_parts.append("UPPER(sf.vehiculo) LIKE ?")
        params.append(f'%{vehiculo_fil.upper()}%')

    where_sql = ('WHERE ' + ' AND '.join(where_parts)) if where_parts else ''

    conn = get_db()
    sugerencias = conn.execute(
        f"""SELECT sf.id, sf.vehiculo, sf.descripcion, sf.estado, sf.timestamp, sf.respuesta_admin,
                   fe.nombre_articulo, fe.tipo_filtro,
                   u.nombre_completo, u.rol
            FROM sugerencias_filtros sf
            JOIN usuarios u ON u.id = sf.usuario_id
            LEFT JOIN filtros_equipo fe ON fe.id = sf.filtro_id
            {where_sql}
            ORDER BY sf.timestamp DESC""",
        params
    ).fetchall()

    stats = conn.execute(
        """SELECT
               COUNT(*) AS total,
               SUM(CASE WHEN estado = 'pendiente'  THEN 1 ELSE 0 END) AS pendientes,
               SUM(CASE WHEN estado = 'revisada'   THEN 1 ELSE 0 END) AS revisadas,
               SUM(CASE WHEN estado = 'aplicada'   THEN 1 ELSE 0 END) AS aplicadas,
               SUM(CASE WHEN estado = 'rechazada'  THEN 1 ELSE 0 END) AS rechazadas
           FROM sugerencias_filtros"""
    ).fetchone()
    conn.close()

    return render_template('admin/sugerencias.html',
        sugerencias=sugerencias,
        estado_fil=estado_fil,
        vehiculo_fil=vehiculo_fil,
        stats=stats,
    )


@app.route('/admin/sugerencias/<int:sug_id>/estado', methods=['POST'])
@admin_required
def admin_sugerencia_estado(sug_id):
    data        = request.get_json(force=True) or {}
    nuevo_estado = data.get('estado', '')
    respuesta   = (data.get('respuesta', '') or '').strip() or None

    if nuevo_estado not in ('revisada', 'aplicada', 'rechazada'):
        return jsonify({'success': False, 'error': 'Estado inválido.'}), 400

    conn = get_db()
    try:
        row = conn.execute(
            "SELECT id FROM sugerencias_filtros WHERE id = ?", (sug_id,)
        ).fetchone()
        if not row:
            return jsonify({'success': False, 'error': 'Sugerencia no encontrada.'}), 404
        conn.execute(
            "UPDATE sugerencias_filtros SET estado = ?, respuesta_admin = ? WHERE id = ?",
            (nuevo_estado, respuesta, sug_id)
        )
        conn.commit()
    finally:
        conn.close()
    return jsonify({'success': True, 'estado': nuevo_estado})


@app.route('/admin/sugerencias/<int:sug_id>/eliminar', methods=['DELETE'])
@admin_required
def admin_sugerencia_eliminar(sug_id):
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT id FROM sugerencias_filtros WHERE id = ?", (sug_id,)
        ).fetchone()
        if not row:
            return jsonify({'success': False, 'error': 'Sugerencia no encontrada.'}), 404
        conn.execute("DELETE FROM sugerencias_filtros WHERE id = ?", (sug_id,))
        conn.commit()
    finally:
        conn.close()
    return jsonify({'success': True})


# ---------------------------------------------------------------------------
# Técnico — equipos en taller
# ---------------------------------------------------------------------------

@app.route('/taller')
@tecnico_required
def taller():
    conn = get_db()
    sync_id = _current_sync_id(conn)

    equipos_taller = []
    familias, categorias = [], []

    if sync_id:
        rows = conn.execute(
            """SELECT e.vehiculo, e.familia, e.categoria,
                      GROUP_CONCAT(DISTINCT e.rutina) AS rutinas,
                      MAX(r.timestamp) AS fecha_ejecucion
               FROM equipos e
               JOIN solicitudes s ON s.equipo_id = e.id AND s.sync_id = e.sync_id
               JOIN respuestas r  ON r.solicitud_id = s.id AND r.accion = 'ejecutado'
               WHERE e.sync_id = ?
               GROUP BY e.vehiculo
               ORDER BY e.vehiculo""",
            (sync_id,)
        ).fetchall()

        equipos_taller = [dict(row) for row in rows]
        familias   = sorted({r['familia']   for r in equipos_taller if r['familia']})
        categorias = sorted({r['categoria'] for r in equipos_taller if r['categoria']})

    conn.close()

    return render_template('tecnico/taller.html',
        equipos=equipos_taller,
        familias=familias,
        categorias=categorias,
        current_sync_id=sync_id,
    )


@app.route('/taller/flota')
@login_required
def taller_flota():
    conn = get_db()
    rows = conn.execute(
        """SELECT equipo, MIN(tipo) AS tipo, COUNT(*) AS num_filtros
           FROM filtros_equipo
           GROUP BY equipo
           ORDER BY equipo"""
    ).fetchall()
    conn.close()
    return render_template('tecnico/flota.html', vehiculos=[dict(r) for r in rows])


# ---------------------------------------------------------------------------
# Admin — eliminar solicitud
# ---------------------------------------------------------------------------

@app.route('/admin/solicitud/<int:solicitud_id>/eliminar', methods=['DELETE'])
@admin_required
def admin_eliminar_solicitud(solicitud_id):
    conn = get_db()
    try:
        sol = conn.execute(
            """SELECT s.id, s.fecha_solicitud,
                      e.vehiculo, e.familia, e.rutina,
                      u.nombre_completo AS solicitado_por_nombre,
                      r.accion, r.comentario_libre, m.descripcion AS motivo_desc
               FROM solicitudes s
               JOIN equipos e   ON e.id = s.equipo_id
               JOIN usuarios u  ON u.id = s.solicitado_por
               LEFT JOIN respuestas r       ON r.solicitud_id = s.id
               LEFT JOIN catalogo_motivos m ON m.id = r.motivo_id
               WHERE s.id = ?""",
            (solicitud_id,)
        ).fetchone()

        if not sol:
            return jsonify({'success': False, 'error': 'Solicitud no encontrada.'}), 404

        rutina_corta = (sol['rutina'][:80] + '…') if sol['rutina'] and len(sol['rutina']) > 80 else (sol['rutina'] or '—')
        detalle = (f'Eliminó solicitud #{solicitud_id}: vehículo={sol["vehiculo"]}, '
                   f'familia={sol["familia"] or "—"}, rutina={rutina_corta}, '
                   f'solicitado_por={sol["solicitado_por_nombre"]}, '
                   f'fecha={sol["fecha_solicitud"]}')
        if sol['accion']:
            detalle += f', respuesta: {sol["accion"]}'
            if sol['motivo_desc']:
                detalle += f' — {sol["motivo_desc"]}'
            if sol['comentario_libre']:
                detalle += f' ({sol["comentario_libre"]})'

        conn.execute("DELETE FROM respuestas WHERE solicitud_id = ?", (solicitud_id,))
        conn.execute("DELETE FROM solicitudes WHERE id = ?", (solicitud_id,))
        _log_actividad(conn, current_user.id, 'eliminar_solicitud', detalle)
        conn.commit()
    finally:
        conn.close()
    return jsonify({'success': True})


# ---------------------------------------------------------------------------
# Admin — auditoría
# ---------------------------------------------------------------------------

@app.route('/admin/auditoria')
@admin_required
def admin_auditoria():
    page        = max(1, request.args.get('page', 1, type=int))
    per_page    = 50
    fecha_desde = request.args.get('fecha_desde', '').strip()
    fecha_hasta = request.args.get('fecha_hasta', '').strip()

    where_parts = ["l.accion_tipo = 'eliminar_solicitud'"]
    params = []
    if fecha_desde:
        where_parts.append("l.timestamp >= ?")
        params.append(fecha_desde)
    if fecha_hasta:
        where_parts.append("l.timestamp <= ?")
        params.append(fecha_hasta + 'T23:59:59')

    where_sql = 'WHERE ' + ' AND '.join(where_parts)

    conn = get_db()
    try:
        total = conn.execute(
            f"SELECT COUNT(*) AS c FROM log_actividad l {where_sql}", params
        ).fetchone()['c']

        offset = (page - 1) * per_page
        logs = conn.execute(
            f"""SELECT l.id, l.timestamp, l.detalle, l.ip_address,
                       COALESCE(u.nombre_completo, 'Sistema') AS nombre_usuario,
                       u.username
                FROM log_actividad l
                LEFT JOIN usuarios u ON u.id = l.usuario_id
                {where_sql}
                ORDER BY l.timestamp DESC
                LIMIT ? OFFSET ?""",
            params + [per_page, offset]
        ).fetchall()
    finally:
        conn.close()

    total_pages = max(1, (total + per_page - 1) // per_page)

    return render_template('admin/auditoria.html',
        logs=logs,
        total=total,
        page=page,
        per_page=per_page,
        total_pages=total_pages,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
    )


@app.route('/admin/log/<int:log_id>/eliminar', methods=['DELETE'])
@admin_required
def admin_eliminar_log(log_id):
    if current_user.rol != 'superadmin':
        return jsonify({'success': False, 'error': 'Sin permiso.'}), 403

    conn = get_db()
    try:
        row = conn.execute("SELECT id FROM log_actividad WHERE id = ?", (log_id,)).fetchone()
        if not row:
            return jsonify({'success': False, 'error': 'Registro no encontrado.'}), 404
        conn.execute("DELETE FROM log_actividad WHERE id = ?", (log_id,))
        conn.commit()
    finally:
        conn.close()
    return jsonify({'success': True})


# ---------------------------------------------------------------------------
# Admin — indicadores — helpers de gráficos SVG
# ---------------------------------------------------------------------------

_DONUT_COLORS = ['#002D6E', '#80AE3F', '#E67E22', '#1E88E5', '#dc2626',
                 '#6b7280', '#8b5cf6', '#06b6d4']

def _build_donut_segments(items, total):
    """Calcula los paths SVG para el gráfico donut de motivos."""
    if not items or not total:
        return []
    cx = cy = 100
    R, r = 88, 52
    segments = []
    angle = -math.pi / 2  # arranca desde el tope
    for i, item in enumerate(items):
        frac = item['cantidad'] / total
        span = 2 * math.pi * min(frac, 0.9999)
        end = angle + span
        large = 1 if span > math.pi else 0
        x1 = round(cx + R * math.cos(angle), 3)
        y1 = round(cy + R * math.sin(angle), 3)
        x2 = round(cx + R * math.cos(end), 3)
        y2 = round(cy + R * math.sin(end), 3)
        x3 = round(cx + r * math.cos(end), 3)
        y3 = round(cy + r * math.sin(end), 3)
        x4 = round(cx + r * math.cos(angle), 3)
        y4 = round(cy + r * math.sin(angle), 3)
        path = (f"M {x1} {y1} A {R} {R} 0 {large} 1 {x2} {y2} "
                f"L {x3} {y3} A {r} {r} 0 {large} 0 {x4} {y4} Z")
        segments.append({
            'path': path,
            'color': _DONUT_COLORS[i % len(_DONUT_COLORS)],
            'label': item['motivo'],
            'count': item['cantidad'],
            'pct': round(frac * 100),
        })
        angle = end
    return segments


def _build_familias_chart_data(familias_cumplimiento):
    """Calcula posiciones SVG para el gráfico de barras horizontales por familia."""
    ROW_H = 36
    LABEL_W = 170
    BAR_W = 320
    PCT_X = LABEL_W + BAR_W + 10
    SVG_W = PCT_X + 46
    rows = []
    for i, f in enumerate(familias_cumplimiento):
        sol = f['solicitados'] or 0
        ejec = f['ejecutados'] or 0
        no_ejec = f['no_ejecutados'] or 0
        pct = round(ejec / sol * 100) if sol else 0
        ejec_px = round((ejec / sol) * BAR_W) if sol else 0
        no_px = round((no_ejec / sol) * BAR_W) if sol else 0
        y = i * ROW_H + 6
        fc = '#80AE3F' if pct >= 80 else ('#E67E22' if pct >= 60 else '#dc2626')
        rows.append({
            'familia': f['familia'],
            'solicitados': sol,
            'ejecutados': ejec,
            'no_ejecutados': no_ejec,
            'pct': pct,
            'ejec_px': ejec_px,
            'no_px': no_px,
            'bar_x': LABEL_W,
            'y': y,
            'y_text': y + ROW_H // 2 + 1,
            'pct_x': PCT_X,
            'pct_color': fc,
        })
    return {
        'rows': rows,
        'svg_h': max(len(rows) * ROW_H + 12, 40),
        'svg_w': SVG_W,
        'label_w': LABEL_W,
        'bar_w': BAR_W,
        'row_h': ROW_H,
    }


# ---------------------------------------------------------------------------
# Admin — indicadores
# ---------------------------------------------------------------------------

@app.route('/admin/indicadores')
@admin_required
def admin_indicadores():
    conn = get_db()
    sync_id = _current_sync_id(conn)

    total_solicitados = ejecutados = no_ejecutados = 0
    pct_ejecucion = 0

    if sync_id:
        row = conn.execute(
            """SELECT COUNT(s.id) AS total,
                      SUM(CASE WHEN r.accion = 'ejecutado'    THEN 1 ELSE 0 END) AS ejec,
                      SUM(CASE WHEN r.accion = 'no_ejecutado' THEN 1 ELSE 0 END) AS no_ej
               FROM solicitudes s
               LEFT JOIN respuestas r ON r.solicitud_id = s.id
               WHERE s.sync_id = ?""",
            (sync_id,)
        ).fetchone()
        total_solicitados = row['total']  or 0
        ejecutados        = row['ejec']   or 0
        no_ejecutados     = row['no_ej']  or 0
        if total_solicitados:
            pct_ejecucion = round(ejecutados / total_solicitados * 100)

    top_motivos = conn.execute(
        """SELECT CASE WHEN r.motivo_id IS NOT NULL
                       THEN COALESCE(m.descripcion, 'Desconocido')
                       ELSE 'Comentario libre'
                  END AS motivo,
                  COUNT(*) AS cantidad
           FROM respuestas r
           LEFT JOIN catalogo_motivos m ON m.id = r.motivo_id
           WHERE r.accion = 'no_ejecutado'
           GROUP BY r.motivo_id
           ORDER BY cantidad DESC
           LIMIT 10"""
    ).fetchall()
    total_no_ej_global = sum(m['cantidad'] for m in top_motivos) or 1

    familias_cumplimiento = conn.execute(
        """SELECT e.familia,
                  COUNT(s.id) AS solicitados,
                  SUM(CASE WHEN r.accion = 'ejecutado'    THEN 1 ELSE 0 END) AS ejecutados,
                  SUM(CASE WHEN r.accion = 'no_ejecutado' THEN 1 ELSE 0 END) AS no_ejecutados
           FROM solicitudes s
           JOIN equipos e ON e.id = s.equipo_id
           LEFT JOIN respuestas r ON r.solicitud_id = s.id
           WHERE e.familia IS NOT NULL
           GROUP BY e.familia
           ORDER BY e.familia"""
    ).fetchall()

    historial_sync = conn.execute(
        """SELECT s.sync_id,
                  MIN(s.fecha_solicitud)                                        AS primera_solicitud,
                  COUNT(s.id)                                                   AS total_solicitados,
                  SUM(CASE WHEN r.accion = 'ejecutado'    THEN 1 ELSE 0 END)  AS ejecutados,
                  SUM(CASE WHEN r.accion = 'no_ejecutado' THEN 1 ELSE 0 END)  AS no_ejecutados
           FROM solicitudes s
           LEFT JOIN respuestas r ON r.solicitud_id = s.id
           GROUP BY s.sync_id
           ORDER BY s.sync_id DESC"""
    ).fetchall()

    donut_segments   = _build_donut_segments(top_motivos, total_no_ej_global)
    familias_chart   = _build_familias_chart_data(familias_cumplimiento)

    conn.close()

    return render_template('admin/indicadores.html',
        total_solicitados=total_solicitados,
        ejecutados=ejecutados,
        no_ejecutados=no_ejecutados,
        pct_ejecucion=pct_ejecucion,
        top_motivos=top_motivos,
        total_no_ej_global=total_no_ej_global,
        familias_cumplimiento=familias_cumplimiento,
        historial_sync=historial_sync,
        current_sync_id=sync_id,
        donut_segments=donut_segments,
        familias_chart=familias_chart,
    )


# ---------------------------------------------------------------------------
# API — búsqueda vehículos (autocomplete navbar)
# ---------------------------------------------------------------------------

@app.route('/api/buscar-vehiculo')
@login_required
def api_buscar_vehiculo():
    q = request.args.get('q', '').strip()
    if len(q) < 2:
        return jsonify([])

    like = f'%{q.upper()}%'
    conn = get_db()

    from_equipos = conn.execute(
        """SELECT DISTINCT vehiculo FROM equipos
           WHERE UPPER(vehiculo) LIKE ?
           ORDER BY vehiculo LIMIT 15""",
        (like,)
    ).fetchall()

    from_filtros = conn.execute(
        """SELECT DISTINCT equipo AS vehiculo FROM filtros_equipo
           WHERE UPPER(equipo) LIKE ?
           ORDER BY equipo LIMIT 15""",
        (like,)
    ).fetchall()

    conn.close()

    seen, results = set(), []
    for r in list(from_equipos) + list(from_filtros):
        v = r['vehiculo']
        if v and v not in seen:
            seen.add(v)
            results.append(v)

    return jsonify(sorted(results)[:12])


# ---------------------------------------------------------------------------
# Ejecuciones No Reportadas
# ---------------------------------------------------------------------------

@app.route('/admin/no-reportadas')
@login_required
def admin_no_reportadas():
    if current_user.rol not in ('admin', 'superadmin', 'cio'):
        flash('Acceso restringido.', 'error')
        return redirect(url_for('dashboard_redirect'))

    estado_fil   = request.args.get('estado', '').strip()
    vehiculo_fil = request.args.get('vehiculo', '').strip()

    where_parts, params = [], []
    if estado_fil:
        where_parts.append("n.estado = ?")
        params.append(estado_fil)
    if vehiculo_fil:
        where_parts.append("UPPER(n.vehiculo) LIKE ?")
        params.append(f'%{vehiculo_fil.upper()}%')

    where_sql = ('WHERE ' + ' AND '.join(where_parts)) if where_parts else ''

    conn = get_db()
    filas = conn.execute(
        f"""SELECT n.id, n.vehiculo, n.familia, n.rutina,
                   n.ind_desviacion_anterior, n.ind_desviacion_nuevo,
                   n.sync_id_anterior, n.sync_id_nuevo,
                   n.estado, n.justificacion, n.timestamp,
                   u.nombre_completo AS registrado_por_nombre
            FROM ejecuciones_no_reportadas n
            LEFT JOIN usuarios u ON u.id = n.registrado_por
            {where_sql}
            ORDER BY n.timestamp DESC""",
        params
    ).fetchall()
    conn.close()

    return render_template('admin/no_reportadas.html',
        filas=filas,
        estado_fil=estado_fil,
        vehiculo_fil=vehiculo_fil,
        es_admin=(current_user.rol in ('admin', 'superadmin')),
    )


@app.route('/admin/no-reportadas/<int:nr_id>/justificar', methods=['POST'])
@admin_required
def admin_no_reportadas_justificar(nr_id):
    data         = request.get_json(force=True) or {}
    estado       = data.get('estado', '')
    justificacion = (data.get('justificacion') or '').strip() or None

    if estado not in ('justificado', 'sin_justificar'):
        return jsonify({'success': False, 'error': 'Estado inválido.'}), 400

    conn = get_db()
    try:
        row = conn.execute(
            "SELECT id FROM ejecuciones_no_reportadas WHERE id = ?", (nr_id,)
        ).fetchone()
        if not row:
            return jsonify({'success': False, 'error': 'Registro no encontrado.'}), 404
        conn.execute(
            """UPDATE ejecuciones_no_reportadas
               SET estado = ?, justificacion = ?, registrado_por = ?
               WHERE id = ?""",
            (estado, justificacion, current_user.id, nr_id)
        )
        conn.commit()
    finally:
        conn.close()
    return jsonify({'success': True, 'estado': estado})


# ---------------------------------------------------------------------------
# Dev
# ---------------------------------------------------------------------------

if __name__ == '__main__':
    os.makedirs('data', exist_ok=True)
    os.makedirs('exports', exist_ok=True)
    app.run(debug=True)
