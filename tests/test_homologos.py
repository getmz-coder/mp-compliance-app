import openpyxl
from models import get_db
from sync_data import sync_homologos


def test_tabla_homologos_existe():
    conn = get_db()
    cols = {row[1] for row in conn.execute("PRAGMA table_info(homologos)")}
    conn.close()
    assert cols == {'id', 'grupo', 'codigo_sap', 'descripcion', 'estado'}


def _make_homo_excel(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Grupos_Homologos'
    ws.append(['Grupo', 'Tamaño Grupo', 'Estado', 'Codigo SAP', 'Descripcion'])
    ws.append(['1', '2', 'Ya correcto', '10045678', 'FILTRO ACEITE MOTOR'])
    ws.append(['1', '2', 'Revisar',     '10045679', 'FILTRO ACEITE ALTERNATIVO'])
    ws.append(['2', '1', 'Ya correcto', '20001234', 'FILTRO HIDRAULICO'])
    wb.save(path)


def test_sync_homologos_inserta(tmp_path):
    path = str(tmp_path / 'homo.xlsx')
    _make_homo_excel(path)

    result = sync_homologos(path)

    assert result['total_registros'] == 3
    assert result['grupos'] == 2

    conn = get_db()
    rows = conn.execute(
        'SELECT * FROM homologos ORDER BY grupo, codigo_sap'
    ).fetchall()
    conn.close()

    assert len(rows) == 3
    assert rows[0]['grupo'] == 1
    assert rows[0]['codigo_sap'] == '10045678'
    assert rows[0]['estado'] == 'Ya correcto'
    assert rows[2]['grupo'] == 2


def test_sync_homologos_reemplaza(tmp_path):
    """Segunda sync borra los anteriores y carga los nuevos."""
    path = str(tmp_path / 'homo.xlsx')
    _make_homo_excel(path)
    sync_homologos(path)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Grupos_Homologos'
    ws.append(['Grupo', 'Tamaño Grupo', 'Estado', 'Codigo SAP', 'Descripcion'])
    ws.append(['5', '1', 'Revisar', '99990001', 'FILTRO NUEVO'])
    wb.save(path)
    result = sync_homologos(path)

    assert result['total_registros'] == 1
    assert result['grupos'] == 1

    conn = get_db()
    count = conn.execute('SELECT COUNT(*) FROM homologos').fetchone()[0]
    conn.close()
    assert count == 1


def test_sync_homologos_columna_faltante(tmp_path):
    """Lanza ValueError si faltan columnas requeridas."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Grupos_Homologos'
    ws.append(['Grupo', 'Estado'])
    path = str(tmp_path / 'bad.xlsx')
    wb.save(path)

    import pytest as _pytest
    with _pytest.raises(ValueError, match='Columnas faltantes'):
        sync_homologos(path)
